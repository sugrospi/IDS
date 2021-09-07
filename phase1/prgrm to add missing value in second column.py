''' to clean the second column we mapped every region with countries and its number of occurence .we found the country
of the missing value by most occuring country in that region basically mod filling'''
def country_finder(a,b):

	
	m=[]
	n=[]
	for i,j in b[a].items():
		m.append(i)
		n.append(j)
	ans=m[n.index(max(n))]
	return ans

	
				
from openpyxl import *
file="1.xlsx"
wb=load_workbook(file)
sheet=wb.active

region_country_size_map=dict()
for i in range(2,100002):
	if sheet.cell(row=i,column=2).value!="####":
		
		if(sheet.cell(row=i,column=1).value not in region_country_size_map ):
			region_country_size_map[sheet.cell(row=i,column=1).value]={}
		if sheet.cell(row=i,column=2).value not in (region_country_size_map[sheet.cell(row=i,column=1).value]):
			(region_country_size_map[sheet.cell(row=i,column=1).value])[sheet.cell(row=i,column=2).value]=0
		(region_country_size_map[sheet.cell(row=i,column=1).value])[sheet.cell(row=i,column=2).value]+=1

for i in range(2,100002):
		if sheet.cell(row=i,column=2).value=="####"  :
			ans=country_finder(sheet.cell(row=i,column=1).value,region_country_size_map)
			sheet.cell(row=i,column=2).value=ans
wb.save(file)
