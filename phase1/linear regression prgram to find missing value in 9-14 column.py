
import winsound	
from openpyxl import *
import numpy as np 
import matplotlib.pyplot as plt 
file="1.xlsx"
wb=load_workbook(file)
sheet=wb.active
def estimate_coef(x, y): 
	# number of observations/points 
	n = np.size(x) 

	# mean of x and y vector 
	m_x, m_y = np.mean(x), np.mean(y) 

	# calculating cross-deviation and deviation about x 
	SS_xy = np.sum((x-m_x)*(y-m_y))
	SS_xx = np.sum((x-m_x)**2)

	# calculating regression coefficients 
	b_1 = SS_xy / SS_xx 
	b_0 = m_y - b_1*m_x 

	return(b_0, b_1) 

def plot_regression_line(x, y, b): 
	# plotting the actual points as scatter plot 
	plt.scatter(x, y, color = "m", 
			marker = "o", s = 30) 

	# predicted response vector 
	y_pred = b[0] + b[1]*x 

	# plotting the regression line 
	plt.plot(x, y_pred, color = "g") 

	# putting labels 
	plt.xlabel('x') 
	plt.ylabel('y') 

	# function to show plot 
	plt.show() 
def answer(x,b):

	return (b[0]+b[1]*x)

def linear_answer(i,j,z):

	x1=[]
	y1=[]

	country=sheet.cell(row=i,column=2).value
	item=sheet.cell(row=i,column=3).value
	w=2
	while(sheet.cell(row=w,column=2).value!=country or sheet.cell(row=w,column=3).value!=item):
		w+=1
	while(sheet.cell(row=w,column=2).value==country and sheet.cell(row=w,column=3).value==item):
		if(sheet.cell(row=w,column=j).value!="####" and sheet.cell(row=w,column=z).value!="####"):
			y1.append(int(sheet.cell(row=w,column=j).value))
			x1.append(int(sheet.cell(row=w,column=z).value))
		w+=1
	x = np.array(x1) 
	y = np.array(y1) 
	b = estimate_coef(x,y) 
	
	ans=int(answer(int(sheet.cell(row=i,column=z).value),b))
	return ans










try:
	
	file="1.xlsx"
	wb=load_workbook(file)
	sheet=wb.active
	for i in range(2,100002):
		if(sheet.cell(row=i,column="column to edit").value=='####'):
			sheet.cell(row=i,column="column to edit").value=linear_answer(i,"column to edit","column with respect to be edited")




	wb.save(file)
	winsound.Beep(2500,1000)
except Exception as e:
	print("error occured")
	print(e)
	winsound.Beep(4500,1000)

