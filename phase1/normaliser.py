
import winsound	
from openpyxl import *

try:
	
	file="1.xlsx"
	wb=load_workbook(file)
	sheet=wb.active

	


	x9=[]
	x10=[]
	x11=[]
	x12=[]
	x13=[]
	x14=[]
	for i in range(2,100002):
		x9.append(sheet.cell(row=i,column=9).value)
	
	
		x10.append(sheet.cell(row=i,column=10).value)


		x11.append(sheet.cell(row=i,column=11).value)
	
	
		x12.append(sheet.cell(row=i,column=12).value)

	
		x13.append(sheet.cell(row=i,column=13).value)

	
		x14.append(sheet.cell(row=i,column=14).value)

	print("first iteration done")
	j=1000
	m9=min(x9)
	m10=min(x10)
	m11=min(x11)
	m12=min(x12)
	m13=min(x13)
	m14=min(x14)

	M9=max(x9)
	M10=max(x10)
	M11=max(x11)
	M12=max(x12)
	M13=max(x13)
	M14=max(x14)

	for i in range(2,100002):
		a9=sheet.cell(row=i,column=9).value
		a10=sheet.cell(row=i,column=10).value
		a11=sheet.cell(row=i,column=11).value
		a12=sheet.cell(row=i,column=12).value
		a13=sheet.cell(row=i,column=13).value
		a14=sheet.cell(row=i,column=14).value

		

		sheet.cell(row=i,column=9).value=((a9-m9)/(M9-m9))
		sheet.cell(row=i,column=10).value=((a10-m10)/(M10-m10))
		sheet.cell(row=i,column=11).value=((a11-m11)/(M11-m11))
		sheet.cell(row=i,column=12).value=((a12-m12)/(M12-m12))
		sheet.cell(row=i,column=13).value=((a13-m13)/(M13-m13))
		sheet.cell(row=i,column=14).value=((a14-m14)/(M14-m14))

		if (i==j):
			print(j,"iteration completed")
			j+=1000





	wb.save(file)
	winsound.Beep(2500,1000)
except Exception as e:
	print("error occured")
	print(e)
	winsound.Beep(2500,1000)
