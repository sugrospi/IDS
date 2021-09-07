''' to clean the first column we maped the country with the region of all the data , later we checked
the country of the missing region and checked found its region using the mapped data_structure and edited the dataset'''

def region_finder(a,b):

	for i,j in b.items():
		for k in j:
			if k==a:
				return i
				
from openpyxl import *
file="1.xlsx"
wb=load_workbook(file)
sheet=wb.active

region_country_map=dict()
for i in range(2,100002):
	if sheet.cell(row=i,column=1).value!="####":
		
		if(sheet.cell(row=i,column=1).value not in region_country_map ):
			region_country_map[sheet.cell(row=i,column=1).value]=set()
		region_country_map[sheet.cell(row=i,column=1).value].add(sheet.cell(row=i,column=2).value)


for i in range(2,100002):
	if sheet.cell(row=i,column=1).value=="####" and sheet.cell(row=i,column=2).value!="####" :
		ans=region_finder(sheet.cell(row=i,column=2).value,region_country_map)
		sheet.cell(row=i,column=1).value=ans

wb.save(file)
