'''to find missing value in the third column we mapped countries with items with number of occurnce of items
then we filled missing value with most occuring  item in that country'''
def item_finder(a,b):

	
	m=[]
	n=[]
	for i,j in b[a].items():
		m.append(i)
		n.append(j)
	ans=m[n.index(max(n))]
	return ans

	
				
from openpyxl import *
file="1.xlsx"
wb=load_workbook(file)
sheet=wb.active

country_item_size_map=dict()
for i in range(2,100002):
	if sheet.cell(row=i,column=3).value!="####":
		
		if(sheet.cell(row=i,column=2).value not in country_item_size_map ):
			country_item_size_map[sheet.cell(row=i,column=2).value]={}
		if sheet.cell(row=i,column=3).value not in (country_item_size_map[sheet.cell(row=i,column=2).value]):
			(country_item_size_map[sheet.cell(row=i,column=2).value])[sheet.cell(row=i,column=3).value]=0
		(country_item_size_map[sheet.cell(row=i,column=2).value])[sheet.cell(row=i,column=3).value]+=1

for i in range(2,100002):
		if sheet.cell(row=i,column=3).value=="####"  :
			ans=item_finder(sheet.cell(row=i,column=2).value,country_item_size_map)
			sheet.cell(row=i,column=3).value=ans
wb.save(file)


