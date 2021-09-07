import numpy as np
import matplotlib.pyplot as plt
import random
from openpyxl import load_workbook
import time


def estimate_coff(data_x1,data_x2,data_x3,data_x4,data_x5,data_y):

    b=0
    w1=1
    w2=1
    w3=2
    w4=5
    w5=2
    time.sleep(1)
    print("starting training ....")
    for j in range(1000):
        for i in range(len(data_x1)):

            ycap=b+w1*data_x1[i]+w2*data_x2[i]+w3*data_x3[i]+w4*data_x4[i]+w5*data_x5[i]
            loss=(ycap-data_y[i])


            b=b-(0.0001*(loss))
            w1=w1-(0.0001*(loss)*data_x1[i])
            w2=w2-(0.0001*(loss)*data_x2[i])
            w3=w3-(0.0001*(loss)*data_x3[i])
            w4=w4-(0.0001*(loss)*data_x4[i])
            w5=w5-(0.0001*(loss)*data_x5[i])

    time.sleep(1)
    print("training..completed")

    return(b,w1,w2,w3,w4,w5)

def rms_error(b,data_x1,data_x2,data_x3,data_x4,data_x5,data_y):
    loss=[]
    for i in range(len(data_x1)):

        predicted_y=b[0]+b[1]*data_x1[i]+b[2]*data_x2[i]+b[3]*data_x3[i]+b[4]*data_x4[i]+b[5]*data_x5[i]

        l=predicted_y-data_y[i]
        loss.append(l**2)
    rms=(sum(loss)/len(loss))**0.5
    return rms



file="data.xlsx"
wb=load_workbook(file)
sheet=wb.active
print("dataset..loaded.......")
data_xtrain1=[]
data_xtrain2=[]
data_xtrain3=[]
data_xtrain4=[]
data_xtrain5=[]
data_ytrain1=[]


data_xtest1=[]
data_xtest2=[]
data_xtest3=[]
data_xtest4=[]
data_xtest5=[]
data_ytest1=[]
for i in range(2,80000):

            data_xtrain1.append((sheet.cell(row=i,column=9).value))
            data_xtrain2.append((sheet.cell(row=i,column=10).value))
            data_xtrain3.append((sheet.cell(row=i,column=11).value))
            data_xtrain4.append((sheet.cell(row=i,column=12).value))
            data_xtrain5.append((sheet.cell(row=i,column=13).value))
            data_ytrain1.append((sheet.cell(row=i,column=14).value))

for i in range(80001,100000-3):

            data_xtest1.append((sheet.cell(row=i,column=9).value))
            data_xtest2.append((sheet.cell(row=i,column=10).value))
            data_xtest3.append((sheet.cell(row=i,column=11).value))
            data_xtest4.append((sheet.cell(row=i,column=12).value))
            data_xtest5.append((sheet.cell(row=i,column=13).value))
            data_ytest1.append((sheet.cell(row=i,column=14).value))


n=int(input("1.train data again \n2.use the trained weights\n"))

if n==1:
    a=open("weights.txt","w")
    b=estimate_coff(data_xtrain1,data_xtrain2,data_xtrain3,data_xtrain4,data_xtrain5,data_ytrain1)
    a.write(str(b[0])+","+str(b[1])+","+str(b[2])+","+str(b[3])+","+str(b[4])+","+str(b[5]))
    r=rms_error(b,data_xtest1,data_xtest2,data_xtest3,data_xtest4,data_xtest5,data_ytest1)
    print("root mean square error:",r)

    print("accuracy:","%.2f" %((1-r)*100),"%")

elif n==2:
    k=open("weights.txt","r")
    if(list(k.read())==[]):
        k.close()
        a=open("weights.txt","w")
        b=estimate_coff(data_xtrain1,data_xtrain2,data_xtrain3,data_xtrain4,data_xtrain5,data_ytrain1)
        a.write(str(b[0])+","+str(b[1])+","+str(b[2])+","+str(b[3])+","+str(b[4])+","+str(b[5]))
        r=rms_error(b,data_xtest1,data_xtest2,data_xtest3,data_xtest4,data_xtest5,data_ytest1)
        print("root mean square error:",r)

        print("accuracy:","%.2f" %((1-r)*100),"%")
    else:
        k.close()
        a=open("weights.txt","r")
        b=(a.read()).split(",")
        b=[float(i) for i in b]
        r=rms_error(b,data_xtest1,data_xtest2,data_xtest3,data_xtest4,data_xtest5,data_ytest1)
        print("root mean square error:",r)

        print("accuracy:","%.2f" %((1-r)*100),"%")



a.close()
